import openpyxl
import os


class Exmake:

    @classmethod
    def chek_init(cls, clmn):
        if type(clmn) != int:
            raise TypeError('Количество столбцов должно быть целым числом')

    @property
    def clmn(self):
        return self._clmn

    @clmn.setter
    def clmn(self, clmn):
        self.chek_init(clmn)
        self._clmn = clmn

    def __init__(self, path, clmn, name):
        self._path = path
        self.clmn = clmn
        self._name = name
        self._files_path = []
        self._cell = []

    def search_files(self):
        ew_search = r'.xlsx'
        for rootdir, dirs, files in os.walk(self._path):
            for file in files:
                if file.endswith(ew_search):
                    self._files_path.append(rootdir + '\\' + file)

    def read_files(self):
        for file in self._files_path:
            book = openpyxl.open(file, read_only=True)
            sheet = book.active
            for i in sheet.iter_rows(max_row=1, min_col=1, max_col=self._clmn, values_only=True):
                self._cell.append(i)

    def write_cell(self):
        book = openpyxl.Workbook()
        sheet = book.active
        count = 1
        for item in self._cell:
            count_column = 1
            for i in item:
                sheet.cell(row=count, column=count_column).value = i
                count_column += 1
            count += 1
        book.save(f'{self._name}.xlsx')
        book.close()


def main():
    path = input('Введите путь до папки: ')
    clmn = int(input('Введите количество столбцов: '))
    name = str(input('Имя файла для сохранения: '))
    make = Exmake(path, clmn, name)
    make.search_files()
    make.read_files()
    make.write_cell()
    root_dir = os.path.dirname(os.path.abspath(__file__))
    print(f'Скрипт отпработал, путь до файла: {root_dir}' + '\\' + getattr(make, '_name') + '.xlsx')


if __name__ == '__main__':
    main()
